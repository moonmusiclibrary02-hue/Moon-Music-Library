from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, status, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response, RedirectResponse, StreamingResponse
from io import BytesIO
from dotenv import load_dotenv
import tempfile
from google.cloud.exceptions import NotFound, Forbidden
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import hashlib
from jose import JWTError, jwt
from datetime import datetime, timedelta, timezone
import os
import logging
import string
import threading
import time
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
import aiofiles
import mimetypes
from bson import ObjectId
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import urllib.request
import re
from urllib.parse import urlparse, parse_qs
from google.cloud import storage
import base64
from google.auth import impersonated_credentials
from google.auth.transport.requests import Request
import google.auth


ROOT_DIR = Path(__file__).parent
load_dotenv()

# Security
SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480  # 8 hours to prevent frequent session expiration

# Simple password hashing with SHA256
security = HTTPBearer()

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Initialize logger
logger = logging.getLogger(__name__)

# --- Google Cloud Storage (GCS) Connection ---
# This environment variable will be injected from Secret Manager by our CI/CD pipeline.
GCS_BUCKET_NAME = os.environ.get('GCS_BUCKET_NAME')
SIGNING_SERVICE_ACCOUNT_EMAIL = os.environ.get('SIGNING_SERVICE_ACCOUNT_EMAIL') # Add this line

# Validate GCS configuration before creating client
if not GCS_BUCKET_NAME:
    logger.critical("FATAL: GCS_BUCKET_NAME environment variable is not set.")
    raise SystemExit(1)

# Initialize GCS client
try:
    storage_client = storage.Client()
    # Verify we can access the bucket (optional)
    storage_client.get_bucket(GCS_BUCKET_NAME)
    logger.info(f"Successfully connected to GCS bucket: {GCS_BUCKET_NAME}")
except Exception as e:
    logger.critical(f"FATAL: Failed to initialize GCS client or access bucket: {str(e)}")
    raise SystemExit(1) from e



# Create the main app
app = FastAPI(title="Music Production Inventory")



# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    user_type: str = "admin"  # "admin" or "manager"
    manager_id: Optional[str] = None  # Link to manager record if user_type is "manager"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str
    email: str
    password: str

class UserLogin(BaseModel):
    email: str
    password: str

class ForgotPasswordRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

class AdminPasswordUpdateRequest(BaseModel):
    user_id: str
    new_password: str
    notify_user: bool = True

class Manager(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    assigned_language: str
    phone: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class ManagerCreate(BaseModel):
    name: str
    email: str
    assigned_language: str
    phone: Optional[str] = None
    custom_password: Optional[str] = None  # Admin can set custom password

class ManagerUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    assigned_language: Optional[str] = None
    phone: Optional[str] = None
    is_active: Optional[bool] = None

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class MusicTrack(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    unique_code: Optional[str] = None  # User-provided unique identifier for the track
    rights_type: Optional[str] = None  # "original" or "multi_rights"
    track_category: Optional[str] = None  # "cover_song" or "original_composition" (only for original tracks)
    rights_details: Optional[str] = None  # "multi_rights" or "own_rights"
    serial_number: Optional[str] = None  # Auto-generated based on rights_type (OC or MR prefix)
    title: str
    music_composer: str
    lyricist: str
    singer_name: str
    tempo: Optional[str] = None
    scale: Optional[str] = None
    audio_language: str
    release_date: Optional[str] = None
    album_name: Optional[str] = None
    other_info: Optional[str] = None
    # GCS blob names for file storage
    mp3_blob_name: Optional[str] = None
    lyrics_blob_name: Optional[str] = None
    session_blob_name: Optional[str] = None
    singer_agreement_blob_name: Optional[str] = None
    music_director_agreement_blob_name: Optional[str] = None
    # Legacy file paths - to be removed after migration
    mp3_file_path: Optional[str] = None
    lyrics_file_path: Optional[str] = None
    session_file_path: Optional[str] = None
    singer_agreement_file_path: Optional[str] = None
    music_director_agreement_file_path: Optional[str] = None
    mp3_filename: Optional[str] = None
    lyrics_filename: Optional[str] = None
    session_filename: Optional[str] = None
    singer_agreement_filename: Optional[str] = None
    music_director_agreement_filename: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    managed_by: Optional[str] = None

class MusicTrackCreate(BaseModel):
    unique_code: str
    rights_type: str  # "original" or "multi_rights"
    track_category: Optional[str] = None  # "cover_song" or "original_composition" (only for original tracks)
    rights_details: Optional[str] = None  # "multi_rights" or "own_rights"
    title: str
    music_composer: str
    lyricist: str
    singer_name: str
    tempo: Optional[str] = None
    scale: Optional[str] = None
    audio_language: str
    release_date: Optional[str] = None
    album_name: Optional[str] = None
    other_info: Optional[str] = None

class MusicTrackUpdate(BaseModel):
    unique_code: Optional[str] = None
    rights_type: Optional[str] = None
    track_category: Optional[str] = None
    rights_details: Optional[str] = None
    title: Optional[str] = None
    music_composer: Optional[str] = None
    lyricist: Optional[str] = None
    singer_name: Optional[str] = None
    tempo: Optional[str] = None
    scale: Optional[str] = None
    audio_language: Optional[str] = None
    release_date: Optional[str] = None
    album_name: Optional[str] = None
    other_info: Optional[str] = None

class BulkUploadResponse(BaseModel):
    successful_count: int
    failed_count: int
    errors: List[dict] = []
    successful_tracks: List[str] = []

# Security functions
def verify_password(plain_password, hashed_password):
    return hashlib.sha256(plain_password.encode()).hexdigest() == hashed_password

def get_password_hash(password):
    return hashlib.sha256(password.encode()).hexdigest()

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            logger.warning("JWT token missing 'sub' claim")
            raise HTTPException(status_code=401, detail="Could not validate credentials")
    except JWTError as e:
        logger.warning(f"JWT decode error: {str(e)}")
        raise HTTPException(status_code=401, detail="Could not validate credentials")
    
    user = await db.users.find_one({"id": user_id})
    if user is None:
        logger.warning(f"User not found for ID: {user_id}")
        raise HTTPException(status_code=401, detail="User not found")
    return User(**user)

# Mock Email Service (replace with real email service later)
async def send_manager_login_credentials(email: str, name: str, password: str, password_source: str = "auto-generated"):
    """
    Mock email service that logs manager login credentials.
    In production, replace this with actual email sending (SendGrid, etc.)
    """
    email_content = f"""
    ═══════════════════════════════════════════════════════
    🎵 MOON MUSIC - Manager Account Created
    ═══════════════════════════════════════════════════════
    
    Hello {name},
    
    Your manager account has been created for Moon Music Inventory System.
    
    🔑 Login Credentials:
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    📧 Email: {email}
    🔒 Password: {password}
    🌐 Login URL: https://music-tracker-6.preview.emergentagent.com
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    📌 Important Notes:
    • Please change your password after first login
    • You can access the Manager Profile to update your information
    • Contact your administrator if you have any issues
    
    Welcome to the team!
    
    ═══════════════════════════════════════════════════════
    """
    
    # Log to server console (visible in supervisorctl logs)
    logger.info("="*60)
    logger.info("📧 MANAGER LOGIN CREDENTIALS SENT")
    logger.info("="*60)
    logger.info(f"Manager: {name}")
    logger.info(f"Email: {email}")
    logger.info(f"Password: {password}")
    logger.info(f"Password Type: {password_source}")
    logger.info(f"Login URL: https://music-tracker-6.preview.emergentagent.com")
    logger.info("="*60)
    
    # Store email log in database for admin reference
    email_log = {
        "id": str(uuid.uuid4()),
        "recipient_email": email,
        "recipient_name": name,
        "subject": "Moon Music - Manager Account Created",
        "content": email_content,
        "credentials": {
            "email": email,
            "password": password,
            "password_source": password_source
        },
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "status": "logged",  # In production, this would be "sent"
        "type": "manager_credentials"
    }
    
    await db.email_logs.insert_one(email_log)
    logger.info(f"📝 Email logged to database with ID: {email_log['id']}")
    
    return True

async def send_password_update_notification(email: str, username: str, new_password: str):
    """
    Send password update notification to user
    """
    email_content = f"""
    ═══════════════════════════════════════════════════════
    🎵 MOON MUSIC - Password Updated
    ═══════════════════════════════════════════════════════
    
    Hello {username},
    
    Your password has been updated by an administrator.
    
    🔑 New Login Credentials:
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    📧 Email: {email}
    🔒 New Password: {new_password}
    🌐 Login URL: https://music-tracker-6.preview.emergentagent.com
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    📌 Important Notes:
    • Please change your password after logging in for security
    • Contact your administrator if you have any issues
    • Keep your password secure and don't share it
    
    Best regards,
    Moon Music Team
    
    ═══════════════════════════════════════════════════════
    """
    
    # Log to server console
    logger.info("="*60)
    logger.info("📧 PASSWORD UPDATE NOTIFICATION")
    logger.info("="*60)
    logger.info(f"User: {username}")
    logger.info(f"Email: {email}")
    logger.info(f"New Password: {new_password}")
    logger.info("="*60)
    
    # Store email log in database
    email_log = {
        "id": str(uuid.uuid4()),
        "recipient_email": email,
        "recipient_name": username,
        "subject": "Moon Music - Password Updated",
        "content": email_content,
        "credentials": {
            "email": email,
            "password": new_password,
            "password_source": "admin-updated"
        },
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "status": "logged",
        "type": "password_update"
    }
    
    await db.email_logs.insert_one(email_log)
    return True

async def send_password_reset_notification(email: str, username: str, new_password: str, admin_name: str):
    """
    Send password reset notification to user
    """
    email_content = f"""
    ═══════════════════════════════════════════════════════
    🎵 MOON MUSIC - Password Reset by Administrator
    ═══════════════════════════════════════════════════════
    
    Hello {username},
    
    Your password has been reset by administrator {admin_name}.
    
    🔑 New Login Credentials:
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    📧 Email: {email}
    🔒 New Password: {new_password}
    🌐 Login URL: https://music-tracker-6.preview.emergentagent.com
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    📌 Important Security Notes:
    • This password was generated automatically for security
    • Please login and change it to your preferred password immediately
    • Never share your password with anyone
    • Contact your administrator if you didn't request this reset
    
    Reset by: {admin_name}
    Time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}
    
    Best regards,
    Moon Music Team
    
    ═══════════════════════════════════════════════════════
    """
    
    # Log to server console
    logger.info("="*60)
    logger.info("📧 PASSWORD RESET NOTIFICATION")
    logger.info("="*60)
    logger.info(f"User: {username}")
    logger.info(f"Email: {email}")
    logger.info(f"New Password: {new_password}")
    logger.info(f"Reset by Admin: {admin_name}")
    logger.info("="*60)
    
    # Store email log in database
    email_log = {
        "id": str(uuid.uuid4()),
        "recipient_email": email,
        "recipient_name": username,
        "subject": "Moon Music - Password Reset",
        "content": email_content,
        "credentials": {
            "email": email,
            "password": new_password,
            "password_source": "admin-reset"
        },
        "reset_by": admin_name,
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "status": "logged",
        "type": "password_reset"
    }
    
    await db.email_logs.insert_one(email_log)
    return True

# Helper functions
def prepare_for_mongo(data):
    """Convert datetime objects to ISO strings for MongoDB storage"""
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, datetime):
                data[key] = value.isoformat()
    return data

def parse_from_mongo(item):
    """Parse datetime strings from MongoDB"""
    if isinstance(item, dict):
        for key, value in item.items():
            if key in ['created_at'] and isinstance(value, str):
                try:
                    item[key] = datetime.fromisoformat(value)
                except:
                    pass
    return item

def extract_google_drive_file_id(url):
    """Extract file ID from Google Drive URL"""
    if not url or not isinstance(url, str):
        return None
    
    # Various Google Drive URL patterns
    patterns = [
        r'/file/d/([a-zA-Z0-9-_]+)',  # Standard sharing URL
        r'id=([a-zA-Z0-9-_]+)',       # URL parameter
        r'drive.google.com/open\?id=([a-zA-Z0-9-_]+)',
        r'drive.google.com/uc\?id=([a-zA-Z0-9-_]+)'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    
    return None

async def download_from_google_drive(url, filename):
    """
    Download file from Google Drive URL to a temporary file.
    
    Args:
        url: Google Drive sharing URL
        filename: Original filename with extension
    
    Returns:
        str: Path to the temporary file with downloaded content
        
    Raises:
        ValueError: If file ID cannot be extracted or download fails
        URLError: If network or server errors occur
        IOError: If file writing fails
    """
    try:
        file_id = extract_google_drive_file_id(url)
        if not file_id:
            raise ValueError("Could not extract file ID from Google Drive URL")
        
        # Use direct download URL
        download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
        
        # Create a temporary file with the correct extension
        suffix = Path(filename).suffix
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temp_file:
            try:
                # Download and write the content to the temp file
                urllib.request.urlretrieve(download_url, temp_file.name)
                return temp_file.name
            except Exception as e:
                # Clean up the temp file if download fails
                try:
                    os.unlink(temp_file.name)
                except Exception as cleanup_exc:
                    logger.warning(f"Failed to clean up temporary file {temp_file.name}: {cleanup_exc}")
                raise
                
    except ValueError as e:
        logger.error(f"Invalid Google Drive URL or file ID: {e}")
        raise
    except urllib.error.URLError as e:
        logger.error(f"Network error downloading from Google Drive: {e}")
        raise ValueError(f"Failed to download file - network error: {str(e)}")
    except IOError as e:
        logger.error(f"IO error while writing downloaded file: {e}")
        raise ValueError(f"Failed to save downloaded file: {str(e)}")
    except Exception as e:
        logger.error(f"Unexpected error downloading from Google Drive: {e}")
        raise ValueError(f"Failed to download file: {str(e)}")

def generate_excel_template():
    """Generate Excel template for bulk upload"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Track Upload"
    
    # Define headers
    headers = [
        'Title*', 'Music Composer*', 'Lyricist*', 'Singer Name*', 
        'Audio Language*', 'Rights Type*', 'Track Category', 
        'Tempo', 'Scale', 'Album Name', 'Release Date', 
        'Other Info', 'Audio File Google Drive Link', 
        'Lyrics File Google Drive Link', 'Session File Google Drive Link',
        'Singer Agreement Google Drive Link', 'Music Director Agreement Google Drive Link'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data row
    sample_data = [
        'My Song Title', 'John Composer', 'Jane Lyricist', 'Singer Name',
        'English', 'original', 'original_composition', 
        '120 BPM', 'C Major', 'My Album', '2024-01-15',
        'Additional song information', 
        'https://drive.google.com/file/d/your_audio_file_id/view',
        'https://drive.google.com/file/d/your_lyrics_file_id/view',
        'https://drive.google.com/file/d/your_session_file_id/view',
        'https://drive.google.com/file/d/your_singer_agreement_id/view',
        'https://drive.google.com/file/d/your_music_director_agreement_id/view'
    ]
    
    for col, data in enumerate(sample_data, 1):
        ws.cell(row=2, column=col, value=data)
    
    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    instructions = [
        "BULK UPLOAD INSTRUCTIONS:",
        "",
        "1. Fill in the required fields marked with * (asterisk)",
        "2. Rights Type must be either 'original' or 'multi_rights'",
        "3. Track Category (only for original tracks): 'original_composition' or 'cover_song'",
        "4. Audio Language should match your assigned language (for managers)",
        "5. Release Date format: YYYY-MM-DD (e.g., 2024-01-15)",
        "",
        "GOOGLE DRIVE LINKS:",
        "- Make sure files are set to 'Anyone with the link can view'",
        "- Use the shareable link from Google Drive",
        "- Audio files should be MP3 format",
        "- Session files should be ZIP or RAR format",
        "- Agreement files should be PDF format",
        "",
        "SUPPORTED FILE TYPES:",
        "- Audio: MP3",
        "- Lyrics: TXT, DOC, DOCX",
        "- Session: ZIP, RAR",
        "- Agreements: PDF",
        "",
        "NOTE: All files will be downloaded from Google Drive during upload process."
    ]
    
    for row, instruction in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=row, column=1, value=instruction)
        if row == 1:
            cell.font = Font(bold=True, size=14)
        elif instruction.endswith(":"):
            cell.font = Font(bold=True)
    
    # Adjust column widths
    for sheet in [ws, instructions_ws]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    return wb

import asyncio

async def upload_to_gcs(file: UploadFile, folder: str) -> str:
    """
    Uploads a file to a specified folder in the GCS bucket and returns its blob name.
    """
    if not GCS_BUCKET_NAME:
        raise HTTPException(status_code=500, detail="GCS bucket name is not configured.")
    
    try:
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        
        # Create a unique filename using a UUID to prevent collisions.
        # Format: audio/123e4567-e89b-12d3-a456-426614174000_my-song.mp3
        blob_name = f"{folder}/{uuid.uuid4()}_{file.filename}"
        blob = bucket.blob(blob_name)
        
        # Stream the file directly to GCS using upload_from_file
        # This avoids loading the entire file into memory
        await asyncio.to_thread(
            blob.upload_from_file,
            file.file,
            content_type=file.content_type,
            rewind=True
        )
        
        logger.info(f"Successfully uploaded {file.filename} to gs://{GCS_BUCKET_NAME}/{blob_name}")
    except Exception as e:
        logger.exception(f"Failed to upload {file.filename} to GCS")
        raise HTTPException(status_code=500, detail=f"Could not upload file: {e}") from e
    
    return blob_name

async def generate_signed_url(blob_name: str, expiration_minutes: int = 60) -> str:
    """
    Generate a signed URL for a GCS blob with configurable expiration time.
    """
    if not GCS_BUCKET_NAME or not blob_name:
        raise HTTPException(status_code=500, detail="GCS bucket name or blob name not configured.")
    
    try:
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(blob_name)
        
        # Generate signed URL that expires in specified minutes
        url = await asyncio.to_thread(
            blob.generate_signed_url,
            expiration=timedelta(minutes=expiration_minutes),
            method="GET",
            version="v4"
        )
        return url
    except Exception as e:
        logger.exception(f"Failed to generate signed URL for blob: {blob_name}")
        raise HTTPException(status_code=500, detail=f"Could not generate signed URL: {e}") from e
async def read_gcs_text(blob_name: str) -> str:
    """
    Read text content from a GCS blob.
    """
    if not GCS_BUCKET_NAME or not blob_name:
        raise HTTPException(status_code=500, detail="GCS bucket name or blob name not configured.")
    
    try:
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(blob_name)
        
        # Download the blob content as text using a thread to avoid blocking
        content = await asyncio.to_thread(blob.download_as_text)
        return content
    except Exception as e:
        logger.exception(f"Failed to read text from blob: {blob_name}")
        raise HTTPException(status_code=500, detail=f"Could not read file content: {e}") from e
# DELETE helper for GCS
async def delete_from_gcs(file_reference: str):
    """
    Deletes a file from the GCS bucket using either a blob name or a public URL.
    
    Args:
        file_reference: Either a GCS blob name or a public URL from storage.googleapis.com
        
    Returns:
        None. Logs success or failure but doesn't raise exceptions.
    """
    if not GCS_BUCKET_NAME or not file_reference:
        logger.warning("GCS_BUCKET_NAME not configured or no file reference provided. Skipping deletion.")
        return

    try:
        # Normalize input - extract blob name from URL if needed
        blob_name = file_reference
        if file_reference.startswith("https://storage.googleapis.com/"):
            prefix = f"https://storage.googleapis.com/{GCS_BUCKET_NAME}/"
            if not file_reference.startswith(prefix):
                logger.error(f"URL '{file_reference}' does not match bucket {GCS_BUCKET_NAME}. Cannot delete.")
                return
            blob_name = file_reference.replace(prefix, "")

        # Get bucket and blob references
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(blob_name)

        # Check existence and delete using asyncio.to_thread for blocking operations
        exists = await asyncio.to_thread(blob.exists)
        if exists:
            await asyncio.to_thread(blob.delete)
            logger.info(f"Successfully deleted gs://{GCS_BUCKET_NAME}/{blob_name}")
        else:
            logger.warning(f"Attempted to delete non-existent blob: {blob_name}")

    except Exception as e:
        logger.exception(f"Failed to delete {blob_name} from GCS")
        # We don't raise an exception as this is often called during cleanup

async def process_bulk_upload_row(row_data, row_number, current_user):
    """Process a single row from bulk upload Excel"""
    try:
        # Extract data from row
        title = str(row_data.get('Title*', '')).strip()
        music_composer = str(row_data.get('Music Composer*', '')).strip()
        lyricist = str(row_data.get('Lyricist*', '')).strip()
        singer_name = str(row_data.get('Singer Name*', '')).strip()
        audio_language = str(row_data.get('Audio Language*', '')).strip()
        rights_type = str(row_data.get('Rights Type*', '')).strip()
        
        # Validate required fields
        required_fields = {
            'Title': title,
            'Music Composer': music_composer,
            'Lyricist': lyricist,
            'Singer Name': singer_name,
            'Audio Language': audio_language,
            'Rights Type': rights_type
        }
        
        missing_fields = [field for field, value in required_fields.items() if not value]
        if missing_fields:
            return None, f"Missing required fields: {', '.join(missing_fields)}"
        
        # Validate rights type
        if rights_type not in ['original', 'multi_rights']:
            return None, f"Invalid rights type '{rights_type}'. Must be 'original' or 'multi_rights'"
        
        # Handle optional fields
        track_category = str(row_data.get('Track Category', '')).strip() or None
        tempo = str(row_data.get('Tempo', '')).strip() or None
        scale = str(row_data.get('Scale', '')).strip() or None
        album_name = str(row_data.get('Album Name', '')).strip() or None
        release_date = str(row_data.get('Release Date', '')).strip() or None
        other_info = str(row_data.get('Other Info', '')).strip() or None
        
        # Validate track category for original tracks
        if rights_type == 'original' and track_category not in ['cover_song', 'original_composition']:
            return None, f"For original tracks, track category must be 'cover_song' or 'original_composition'"
        
        # For managers, validate language
        if current_user.user_type == "manager" and current_user.manager_id:
            manager_record = await db.managers.find_one({"id": current_user.manager_id})
            if manager_record and manager_record.get("assigned_language"):
                assigned_language = manager_record["assigned_language"]
                if audio_language != assigned_language:
                    return None, f"You can only upload tracks in your assigned language: {assigned_language}"
        
        # Generate unique code and serial number
        # Get language code (first 3 characters, uppercase)
        language_code = audio_language[:3].upper()
        
        # Determine prefix based on rights type and category
        if rights_type == "original":
            if track_category == "original_composition":
                prefix = f"{language_code}-OC"
            else:  # cover_song
                prefix = f"{language_code}-OCC"
        else:  # multi_rights
            prefix = f"{language_code}-MR"
        
        # Get the next number for this prefix
        last_track = await db.tracks.find_one(
            {"unique_code": {"$regex": f"^{prefix}"}},
            sort=[("unique_code", -1)]
        )
        
        if last_track and last_track.get("unique_code"):
            try:
                last_number = int(last_track["unique_code"].split(prefix)[1])
                next_number = last_number + 1
            except (ValueError, IndexError):
                next_number = 1
        else:
            next_number = 1
        
        unique_code = f"{prefix}{next_number:04d}"
        
        # Generate serial number
        serial_prefix = "OC" if rights_type == "original" else "MR"
        last_serial_track = await db.tracks.find_one(
            {"serial_number": {"$regex": f"^{serial_prefix}"}},
            sort=[("serial_number", -1)]
        )
        
        if last_serial_track and last_serial_track.get("serial_number"):
            try:
                last_serial_number = int(last_serial_track["serial_number"][2:])
                next_serial_number = last_serial_number + 1
            except (ValueError, IndexError):
                next_serial_number = 1
        else:
            next_serial_number = 1
        
        serial_number = f"{serial_prefix}{next_serial_number:04d}"
        
        # Handle file downloads from Google Drive and upload to GCS
        blob_names = {}
        file_names = {}
        
        google_drive_fields = {
            'Audio File Google Drive Link': ('mp3_blob_name', 'mp3_filename', '.mp3', 'audio'),
            'Lyrics File Google Drive Link': ('lyrics_blob_name', 'lyrics_filename', '.txt', 'lyrics'),
            'Session File Google Drive Link': ('session_blob_name', 'session_filename', '.zip', 'sessions'),
            'Singer Agreement Google Drive Link': ('singer_agreement_blob_name', 'singer_agreement_filename', '.pdf', 'agreements'),
            'Music Director Agreement Google Drive Link': ('music_director_agreement_blob_name', 'music_director_agreement_filename', '.pdf', 'agreements')
        }
        
        for field_name, (blob_key, filename_key, extension, folder) in google_drive_fields.items():
            google_drive_url = str(row_data.get(field_name, '')).strip()
            if google_drive_url and google_drive_url.lower() != 'nan':
                try:
                    temp_filename = f"{uuid.uuid4()}{extension}"
                    temp_file_path = await download_from_google_drive(google_drive_url, temp_filename)
                    
                    # Upload to GCS from the temporary file
                    async with aiofiles.open(temp_file_path, 'rb') as f:
                        content = await f.read()
                        file = UploadFile(
                            filename=temp_filename,
                            file=BytesIO(content),
                            content_type=mimetypes.guess_type(temp_filename)[0] or 'application/octet-stream'
                        )
                        blob_name = await upload_to_gcs(file, folder)
                    
                    # Store the blob name and original filename
                    blob_names[blob_key] = blob_name
                    file_names[filename_key] = temp_filename
                    
                    # Clean up temporary file
                    try:
                        os.remove(temp_file_path)
                    except:
                        pass
                except Exception as e:
                    return None, f"Error processing {field_name}: {str(e)}"
        
        # Set managed_by for managers
        managed_by = None
        if current_user.user_type == "manager" and current_user.manager_id:
            managed_by = current_user.manager_id
        
        # Create track
        track = MusicTrack(
            unique_code=unique_code,
            rights_type=rights_type,
            track_category=track_category,
            serial_number=serial_number,
            title=title,
            music_composer=music_composer,
            lyricist=lyricist,
            singer_name=singer_name,
            tempo=tempo,
            scale=scale,
            audio_language=audio_language,
            release_date=release_date,
            album_name=album_name,
            other_info=other_info,
            created_by=current_user.id,
            managed_by=managed_by,
            **blob_names,  # Store GCS blob names
            **file_names   # Store original filenames
        )
        
        track_dict = prepare_for_mongo(track.dict())
        await db.tracks.insert_one(track_dict)
        
        return track.id, None
        
    except Exception as e:
        logger.error(f"Error processing row {row_number}: {e}")
        return None, f"Unexpected error: {str(e)}"

# Rate limiter setup
class RateLimiter:
    """Thread-safe in-memory rate limiter.
    
    WARNING: This implementation is suitable for development and single-process deployments only.
    For production/multi-process deployments, replace with a distributed rate limiter using Redis
    or similar to ensure limits are enforced across all workers.
    
    Example Redis implementation:
    https://redis.io/commands/incr
    """
    
    def __init__(self, max_requests: int, time_window: int, max_users: int = 10000):
        self.max_requests = max_requests
        self.time_window = time_window  # in seconds
        self.max_users = max_users  # Threshold for cleanup
        self.requests = {}
        self.lock = threading.Lock()
    
    def _cleanup_old_requests(self, now: float, user_id: Optional[str] = None) -> None:
        """Remove expired requests for given user or all users if dict too large"""
        if user_id is not None:
            # Clean specific user's requests
            requests = self.requests.get(user_id, [])
            active = [ts for ts in requests if now - ts < self.time_window]
            if active:
                self.requests[user_id] = active
            else:
                self.requests.pop(user_id, None)
        elif len(self.requests) > self.max_users:
            # Bulk cleanup when too many users
            active_users = {}
            for uid, timestamps in self.requests.items():
                active = [ts for ts in timestamps if now - ts < self.time_window]
                if active:
                    active_users[uid] = active
            self.requests = active_users
    
    def is_allowed(self, user_id: str) -> bool:
        now = time.time()
        
        with self.lock:
            # Clean expired requests for this user
            self._cleanup_old_requests(now, user_id)
            
            # Check rate limit
            user_requests = self.requests.get(user_id, [])
            
            if len(user_requests) >= self.max_requests:
                return False
            
            # Record new request
            user_requests.append(now)
            self.requests[user_id] = user_requests
            
            # Periodic cleanup of all users when threshold reached
            if len(self.requests) > self.max_users:
                self._cleanup_old_requests(now)
            
            return True

# Initialize rate limiter (15 requests per minute per user, cleanup at 10k users)
upload_rate_limiter = RateLimiter(max_requests=15, time_window=60, max_users=10000)

async def require_upload_permission(folder: str, current_user: User):
    """Validate user's permission to upload to specific folder"""
    if current_user.user_type == "admin":
        return  # Admins have full access
    
    if current_user.user_type != "manager":
        raise HTTPException(
            status_code=403,
            detail="Only managers and admins can upload files"
        )
    
    # Check manager exists and is active before any folder access
    if current_user.manager_id:
        manager = await db.managers.find_one({"id": current_user.manager_id})
        if not manager or not manager.get("is_active"):
            raise HTTPException(
                status_code=403,
                detail="Manager account is inactive"
            )
        
        # Folder-specific access control can be added here if needed
        # Example: Restrict manager access to specific folders if required
        # if folder == "audio" and not manager.get("has_audio_access", False):
        #     raise HTTPException(
        #         status_code=403,
        #         detail="Access to audio folder is not permitted for this manager"
        #     )

def sanitize_filename(filename: str) -> str:
    """Sanitize filename for safe storage"""
    # Get basename and split into name and extension
    filename = os.path.basename(filename)
    name, ext = os.path.splitext(filename)
    
    # Define safe characters
    name_safe_chars = set(string.ascii_letters + string.digits + '-_')
    ext_safe_chars = set(string.ascii_letters + string.digits + '.')
    
    # Clean and validate name part
    name = name.strip('.')  # Remove leading/trailing dots
    name = ''.join(c if c in name_safe_chars else '_' for c in name)
    name = '_'.join(filter(None, name.split('.')))  # Collapse multiple dots
    if not name:
        name = 'unnamed'
    
    # Clean and validate extension
    ext = ext.lstrip('.')  # Remove leading dot
    if ext:
        ext = ''.join(c if c in ext_safe_chars else '_' for c in ext)[:10]  # Limit ext length
        ext = '.' + ext if ext else ''  # Add back one dot if we have an extension
    
    # Enforce total length limit while preserving extension
    if len(name) + len(ext) > 255:
        name = name[:255 - len(ext)]
        
    return name + ext

# Routes
@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate):
    # Check if user exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    hashed_password = get_password_hash(user_data.password)
    user = User(
        username=user_data.username,
        email=user_data.email
    )
    
    user_dict = user.dict()
    user_dict["hashed_password"] = hashed_password
    user_dict = prepare_for_mongo(user_dict)
    
    await db.users.insert_one(user_dict)
    return user

@api_router.post("/auth/login", response_model=Token)
async def login(user_data: UserLogin):
    user = await db.users.find_one({"email": user_data.email})
    if not user or not verify_password(user_data.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Incorrect email or password")
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["id"]}, expires_delta=access_token_expires
    )
    
    user_obj = User(**parse_from_mongo(user))
    return Token(access_token=access_token, token_type="bearer", user=user_obj)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    user = await db.users.find_one({"email": request.email})
    if not user:
        # Don't reveal if user exists or not for security
        return {"message": "If an account with that email exists, a password reset link has been sent."}
    
    # Generate reset token (valid for 1 hour)
    reset_token = create_access_token(
        data={"sub": user["id"], "type": "password_reset"}, 
        expires_delta=timedelta(hours=1)
    )
    
    # In production, you would send this token via email
    # For demo purposes, we'll return it in the response
    return {
        "message": "If an account with that email exists, a password reset link has been sent.",
        "reset_token": reset_token,  # In production, this would be sent via email
        "instructions": "Use this token to reset your password. In production, this would be sent to your email."
    }

@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    try:
        # Verify the reset token
        payload = jwt.decode(request.token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        token_type: str = payload.get("type")
        
        if token_type != "password_reset":
            raise HTTPException(status_code=400, detail="Invalid token type")
            
        if user_id is None:
            raise HTTPException(status_code=400, detail="Invalid token")
            
    except JWTError:
        raise HTTPException(status_code=400, detail="Invalid or expired token")
    
    # Update user's password
    hashed_password = get_password_hash(request.new_password)
    result = await db.users.update_one(
        {"id": user_id}, 
        {"$set": {"hashed_password": hashed_password}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Password successfully reset"}

@api_router.post("/managers", response_model=Manager)
async def create_manager(manager_data: ManagerCreate, current_user: User = Depends(get_current_user)):
    # Only admins can create managers
    if current_user.user_type != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if manager with same email exists
    existing_manager = await db.managers.find_one({"email": manager_data.email})
    if existing_manager:
        raise HTTPException(status_code=400, detail="Manager with this email already exists")
    
    # Check if user with same email exists
    existing_user = await db.users.find_one({"email": manager_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="User with this email already exists")
    
    # Create manager record
    manager = Manager(
        name=manager_data.name,
        email=manager_data.email,
        assigned_language=manager_data.assigned_language,
        phone=manager_data.phone,
        created_by=current_user.id
    )
    
    manager_dict = prepare_for_mongo(manager.dict())
    await db.managers.insert_one(manager_dict)
    
    # Use custom password if provided, otherwise generate random password
    if manager_data.custom_password and manager_data.custom_password.strip():
        temp_password = manager_data.custom_password.strip()
        password_source = "admin-set"
    else:
        import secrets
        import string
        temp_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))
        password_source = "auto-generated"
    
    # Create user account for manager
    manager_user = User(
        username=manager_data.name,
        email=manager_data.email,
        user_type="manager",
        manager_id=manager.id
    )
    
    user_dict = manager_user.dict()
    user_dict["hashed_password"] = get_password_hash(temp_password)
    user_dict = prepare_for_mongo(user_dict)
    
    await db.users.insert_one(user_dict)
    
    # Mock email service - log credentials and store in database
    await send_manager_login_credentials(manager_data.email, manager_data.name, temp_password, password_source)
    
    return manager

@api_router.get("/managers", response_model=List[Manager])
async def get_managers(current_user: User = Depends(get_current_user)):
    # Only admins can view managers list
    if current_user.user_type != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    managers = await db.managers.find({"is_active": True}).to_list(1000)
    return [Manager(**parse_from_mongo(manager)) for manager in managers]

@api_router.get("/managers/{manager_id}", response_model=Manager)
async def get_manager(manager_id: str, current_user: User = Depends(get_current_user)):
    # Only admins can view individual manager details
    if current_user.user_type != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
        
    manager = await db.managers.find_one({"id": manager_id, "is_active": True})
    if not manager:
        raise HTTPException(status_code=404, detail="Manager not found")
    return Manager(**parse_from_mongo(manager))

@api_router.put("/managers/{manager_id}", response_model=Manager)
async def update_manager(
    manager_id: str, 
    manager_update: ManagerUpdate, 
    current_user: User = Depends(get_current_user)
):
    # Only admins can update managers
    if current_user.user_type != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
        
    manager = await db.managers.find_one({"id": manager_id, "is_active": True})
    if not manager:
        raise HTTPException(status_code=404, detail="Manager not found")
    
    update_data = {k: v for k, v in manager_update.dict().items() if v is not None}
    
    if update_data:
        await db.managers.update_one({"id": manager_id}, {"$set": update_data})
    
    updated_manager = await db.managers.find_one({"id": manager_id})
    return Manager(**parse_from_mongo(updated_manager))

@api_router.delete("/managers/{manager_id}")
async def delete_manager(manager_id: str, current_user: User = Depends(get_current_user)):
    # Only admins can delete managers
    if current_user.user_type != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
        
    manager = await db.managers.find_one({"id": manager_id, "is_active": True})
    if not manager:
        raise HTTPException(status_code=404, detail="Manager not found")
    
    # Soft delete - set is_active to False
    await db.managers.update_one({"id": manager_id}, {"$set": {"is_active": False}})
    return {"message": "Manager deleted successfully"}

@api_router.get("/profile", response_model=dict)
async def get_profile(current_user: User = Depends(get_current_user)):
    profile = {
        "id": current_user.id,
        "username": current_user.username,
        "email": current_user.email,
        "user_type": current_user.user_type,
        "created_at": current_user.created_at
    }
    
    # If user is a manager, include manager details
    if current_user.user_type == "manager" and current_user.manager_id:
        manager = await db.managers.find_one({"id": current_user.manager_id})
        if manager:
            profile["manager_details"] = {
                "name": manager["name"],
                "assigned_language": manager["assigned_language"],
                "phone": manager.get("phone"),
                "is_active": manager.get("is_active", True)
            }
    
    return profile

@api_router.put("/profile/password")
async def update_password(
    old_password: str = Form(...),
    new_password: str = Form(...),
    current_user: User = Depends(get_current_user)
):
    # Get current user from database
    user = await db.users.find_one({"id": current_user.id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify old password
    if not verify_password(old_password, user["hashed_password"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Update password
    hashed_new_password = get_password_hash(new_password)
    await db.users.update_one(
        {"id": current_user.id}, 
        {"$set": {"hashed_password": hashed_new_password}}
    )
    
    return {"message": "Password updated successfully"}

@api_router.put("/admin/users/{user_id}/password")
async def admin_update_user_password(
    user_id: str,
    password_data: AdminPasswordUpdateRequest,
    current_user: User = Depends(get_current_user)
):
    """Admin endpoint to update any user's password"""
    if current_user.user_type != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Find the target user
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update password
    hashed_new_password = get_password_hash(password_data.new_password)
    await db.users.update_one(
        {"id": user_id}, 
        {"$set": {"hashed_password": hashed_new_password}}
    )
    
    # Send notification if requested
    if password_data.notify_user:
        await send_password_update_notification(
            target_user["email"], 
            target_user["username"], 
            password_data.new_password
        )
    
    return {
        "message": f"Password updated successfully for user {target_user['username']}",
        "notification_sent": password_data.notify_user
    }

@api_router.get("/admin/users")
async def get_all_users(current_user: User = Depends(get_current_user)):
    """Admin endpoint to get all users for password management"""
    if current_user.user_type != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = await db.users.find({}, {"hashed_password": 0}).to_list(1000)
    return [User(**parse_from_mongo(user)) for user in users]

@api_router.post("/admin/users/{user_id}/reset-password")
async def admin_reset_user_password(
    user_id: str,
    current_user: User = Depends(get_current_user)
):
    """Admin endpoint to reset user password and send new credentials"""
    if current_user.user_type != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Find the target user
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Generate new random password
    import secrets
    import string
    new_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))
    
    # Update password
    hashed_new_password = get_password_hash(new_password)
    await db.users.update_one(
        {"id": user_id}, 
        {"$set": {"hashed_password": hashed_new_password}}
    )
    
    # Send new credentials
    await send_password_reset_notification(
        target_user["email"], 
        target_user["username"], 
        new_password,
        current_user.username
    )
    
    return {
        "message": f"Password reset successfully for user {target_user['username']}",
        "new_password": new_password,  # For admin reference
        "notification_sent": True
    }

@api_router.get("/admin/email-logs")
async def get_email_logs(current_user: User = Depends(get_current_user)):
    """Get email logs for administrators to see sent credentials"""
    if current_user.user_type != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    email_logs = await db.email_logs.find(
        {"type": "manager_credentials"}, 
        {"_id": 0}
    ).sort("sent_at", -1).limit(50).to_list(50)
    
    return email_logs

@api_router.get("/admin/manager-credentials/{manager_id}")
async def get_manager_credentials(manager_id: str, current_user: User = Depends(get_current_user)):
    """Get specific manager's login credentials for admin"""
    if current_user.user_type != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    manager = await db.managers.find_one({"id": manager_id})
    if not manager:
        raise HTTPException(status_code=404, detail="Manager not found")
    
    # Find the latest email log for this manager
    email_log = await db.email_logs.find_one(
        {"recipient_email": manager["email"], "type": "manager_credentials"},
        sort=[("sent_at", -1)]
    )
    
    if not email_log:
        raise HTTPException(status_code=404, detail="No credentials found for this manager")
    
    return {
        "manager_name": manager["name"],
        "email": email_log["credentials"]["email"],
        "password": email_log["credentials"]["password"],
        "sent_at": email_log["sent_at"],
        "login_url": "https://music-tracker-6.preview.emergentagent.com"
    }

@api_router.put("/profile")
async def update_profile(
    username: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user)
):
    update_data = {}
    
    if username:
        update_data["username"] = username
    
    # Update user record
    if update_data:
        await db.users.update_one({"id": current_user.id}, {"$set": update_data})
    
    # If user is a manager, also update manager record
    if current_user.user_type == "manager" and current_user.manager_id:
        manager_update = {}
        if username:
            manager_update["name"] = username
        if phone:
            manager_update["phone"] = phone
        
        if manager_update:
            await db.managers.update_one(
                {"id": current_user.manager_id}, 
                {"$set": manager_update}
            )
    
    return {"message": "Profile updated successfully"}

@api_router.post("/tracks", response_model=MusicTrack)
async def create_track(
    unique_code: Optional[str] = Form(None),
    rights_type: str = Form(...),
    track_category: Optional[str] = Form(None),
    rights_details: Optional[str] = Form(None),
    title: str = Form(...),
    music_composer: str = Form(...),
    lyricist: str = Form(...),
    singer_name: str = Form(...),
    tempo: Optional[str] = Form(None),
    scale: Optional[str] = Form(None),
    audio_language: str = Form(...),
    release_date: Optional[str] = Form(None),
    album_name: Optional[str] = Form(None),
    other_info: Optional[str] = Form(None),
    managed_by: Optional[str] = Form(None),
    # GCS blob names and filenames from frontend
    mp3_blob_name: Optional[str] = Form(None),
    mp3_filename: Optional[str] = Form(None),
    lyrics_blob_name: Optional[str] = Form(None),
    lyrics_filename: Optional[str] = Form(None),
    session_blob_name: Optional[str] = Form(None),
    session_filename: Optional[str] = Form(None),
    singer_agreement_blob_name: Optional[str] = Form(None),
    singer_agreement_filename: Optional[str] = Form(None),
    music_director_agreement_blob_name: Optional[str] = Form(None),
    music_director_agreement_filename: Optional[str] = Form(None),
    # Legacy: Still accept file uploads for backward compatibility
    mp3_file: Optional[UploadFile] = File(None),
    lyrics_file: Optional[UploadFile] = File(None),
    session_file: Optional[UploadFile] = File(None),
    singer_agreement_file: Optional[UploadFile] = File(None),
    music_director_agreement_file: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user)
):
    # Validate rights type and track category
    if rights_type not in ["original", "multi_rights"]:
        raise HTTPException(status_code=400, detail="Rights type must be 'original' or 'multi_rights'")
    
    if rights_type == "original" and track_category not in ["cover_song", "original_composition"]:
        raise HTTPException(status_code=400, detail="For original tracks, track category must be 'cover_song' or 'original_composition'")
    
    if rights_type == "multi_rights" and track_category is not None:
        raise HTTPException(status_code=400, detail="Track category should not be specified for multi_rights tracks")
    
    # Validate rights details
    if rights_details is not None and rights_details not in ["multi_rights", "own_rights"]:
        raise HTTPException(status_code=400, detail="Rights details must be 'multi_rights' or 'own_rights'")
    
    # Auto-generate unique code if not provided
    if not unique_code:
        # Generate language code (first 3 letters)
        language_code = audio_language[:3].upper() if audio_language else "UNK"
        
        # Determine prefix based on rights type and category
        if rights_type == "original":
            if track_category == "cover_song":
                code_prefix = "OCC"
            elif track_category == "original_composition":
                code_prefix = "OCW"
            else:
                code_prefix = "OC"
        else:  # multi_rights
            code_prefix = "MR"
        
        # Find the next available number for this language-prefix combination
        full_prefix = f"{language_code}-{code_prefix}"
        regex_pattern = f"^{language_code}-{code_prefix}\\d+$"
        
        last_track_with_code = await db.tracks.find_one(
            {"unique_code": {"$regex": regex_pattern}},
            sort=[("unique_code", -1)]
        )
        
        if last_track_with_code and last_track_with_code.get("unique_code"):
            try:
                # Extract number from last unique code
                last_code = last_track_with_code["unique_code"]
                number_part = last_code[len(f"{language_code}-{code_prefix}"):]
                last_number = int(number_part)
                next_number = last_number + 1
            except (ValueError, IndexError):
                next_number = 1
        else:
            next_number = 1
        
        # Generate the unique code
        unique_code = f"{language_code}-{code_prefix}{next_number:04d}"
        
        # Double check this code doesn't exist
        existing_track = await db.tracks.find_one({"unique_code": unique_code})
        if existing_track:
            next_number += 1
            unique_code = f"{language_code}-{code_prefix}{next_number:04d}"
    
    # Generate serial number based on rights type
    prefix = "OC" if rights_type == "original" else "MR"
    
    # Get the next serial number for this prefix
    last_track = await db.tracks.find_one(
        {"serial_number": {"$regex": f"^{prefix}"}},
        sort=[("serial_number", -1)]
    )
    
    if last_track and last_track.get("serial_number"):
        # Extract number from last serial number
        try:
            last_number = int(last_track["serial_number"][2:])  # Remove prefix and convert to int
            next_number = last_number + 1
        except (ValueError, IndexError):
            next_number = 1
    else:
        next_number = 1
    
    serial_number = f"{prefix}{next_number:04d}"  # Format as OC0001, MR0001, etc.
    
    # Log what we received
    logger.info(f"Creating track '{title}' for user {current_user.id}")
    logger.info(f"Received blob names: mp3={mp3_blob_name}, lyrics={lyrics_blob_name}, session={session_blob_name}")
    logger.info(f"Received files: mp3={mp3_file is not None}, lyrics={lyrics_file is not None}, session={session_file is not None}")
    
    # Handle file uploads (legacy) OR use provided blob names (new GCS workflow)
    # Priority: If files are uploaded directly, use them (legacy)
    # Otherwise, use blob names provided from frontend (new workflow)
    
    # MP3 Audio
    if mp3_file:
        # Legacy: Upload file directly
        mp3_blob_name = await upload_to_gcs(mp3_file, "audio")
        mp3_filename = mp3_file.filename
    # else: use mp3_blob_name and mp3_filename from Form parameters (already set)
    
    # Lyrics
    if lyrics_file:
        lyrics_blob_name = await upload_to_gcs(lyrics_file, "lyrics")
        lyrics_filename = lyrics_file.filename
    # else: use lyrics_blob_name and lyrics_filename from Form parameters
        
    # Session
    if session_file:
        session_blob_name = await upload_to_gcs(session_file, "sessions")
        session_filename = session_file.filename
    # else: use session_blob_name and session_filename from Form parameters
        
    # Singer Agreement
    if singer_agreement_file:
        singer_agreement_blob_name = await upload_to_gcs(singer_agreement_file, "agreements")
        singer_agreement_filename = singer_agreement_file.filename
    # else: use singer_agreement_blob_name and singer_agreement_filename from Form parameters
        
    # Music Director Agreement
    if music_director_agreement_file:
        music_director_agreement_blob_name = await upload_to_gcs(music_director_agreement_file, "agreements")
        music_director_agreement_filename = music_director_agreement_file.filename
    # else: use music_director_agreement_blob_name and music_director_agreement_filename from Form parameters
    
    # For managers, validate that they're uploading in their assigned language
    if current_user.user_type == "manager" and current_user.manager_id:
        manager_record = await db.managers.find_one({"id": current_user.manager_id})
        if manager_record and manager_record.get("assigned_language"):
            assigned_language = manager_record["assigned_language"]
            if audio_language != assigned_language:
                raise HTTPException(
                    status_code=400, 
                    detail=f"You can only upload tracks in your assigned language: {assigned_language}"
                )
        # Auto-set managed_by for manager uploads
        managed_by = current_user.manager_id

    # Create track
    track = MusicTrack(
        unique_code=unique_code,
        rights_type=rights_type,
        track_category=track_category,
        rights_details=rights_details,
        serial_number=serial_number,
        title=title,
        music_composer=music_composer,
        lyricist=lyricist,
        singer_name=singer_name,
        tempo=tempo,
        scale=scale,
        audio_language=audio_language,
        release_date=release_date,
        album_name=album_name,
        other_info=other_info,
        # Store GCS blob names
        mp3_blob_name=mp3_blob_name,
        lyrics_blob_name=lyrics_blob_name,
        session_blob_name=session_blob_name,
        singer_agreement_blob_name=singer_agreement_blob_name,
        music_director_agreement_blob_name=music_director_agreement_blob_name,
        # Store original filenames
        mp3_filename=mp3_filename,
        lyrics_filename=lyrics_filename,
        session_filename=session_filename,
        singer_agreement_filename=singer_agreement_filename,
        music_director_agreement_filename=music_director_agreement_filename,
        created_by=current_user.id,
        managed_by=managed_by
    )
    
    track_dict = prepare_for_mongo(track.dict())
    await db.tracks.insert_one(track_dict)
    
    return track

@api_router.get("/tracks", response_model=List[MusicTrack])
async def get_tracks(
    search: Optional[str] = None,
    composer: Optional[str] = None,
    singer: Optional[str] = None,
    album: Optional[str] = None,
    language: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    
    # Role-based filtering: Managers can only see their own tracks
    if current_user.user_type == "manager":
        query["$or"] = [
            {"created_by": current_user.id},  # Tracks created by this manager
            {"managed_by": current_user.manager_id}  # Tracks assigned to this manager
        ]
    # Admins can see all tracks (no additional filter)
    
    if search:
        search_filter = {
            "$or": [
                {"unique_code": {"$regex": search, "$options": "i"}},
                {"title": {"$regex": search, "$options": "i"}},
                {"music_composer": {"$regex": search, "$options": "i"}},
                {"singer_name": {"$regex": search, "$options": "i"}},
                {"album_name": {"$regex": search, "$options": "i"}}
            ]
        }
        if "$or" in query:
            # Combine role filter with search filter
            query = {"$and": [query, search_filter]}
        else:
            query.update(search_filter)
    
    if composer:
        query["music_composer"] = {"$regex": composer, "$options": "i"}
    
    if singer:
        query["singer_name"] = {"$regex": singer, "$options": "i"}
    
    if album:
        query["album_name"] = {"$regex": album, "$options": "i"}
    
    if language:
        query["audio_language"] = {"$regex": language, "$options": "i"}
    
    tracks = await db.tracks.find(query).to_list(1000)
    return [MusicTrack(**parse_from_mongo(track)) for track in tracks]

# REPLACE THE ENTIRE delete_track FUNCTION WITH THIS
@api_router.delete("/tracks/{track_id}")
async def delete_track(track_id: str, current_user: User = Depends(get_current_user)):
    # Step 1: Find the track metadata in MongoDB
    track = await db.tracks.find_one({"id": track_id})
    if not track:
        raise HTTPException(status_code=404, detail="Track not found")

    # Step 2: Authorization (This logic remains unchanged)
    if current_user.user_type == "manager":
        if track.get("created_by") != current_user.id and track.get("managed_by") != current_user.manager_id:
            raise HTTPException(status_code=403, detail="Not authorized to delete this track")

    # Step 3: Delete all associated files from Google Cloud Storage
    blob_name_fields = [
        "mp3_blob_name",
        "lyrics_blob_name",
        "session_blob_name",
        "singer_agreement_blob_name",
        "music_director_agreement_blob_name"
    ]

    # Delete GCS blobs
    for field in blob_name_fields:
        blob_name = track.get(field)
        if blob_name:
            await delete_from_gcs(blob_name)

    # Step 4: After deleting files, delete the record from MongoDB
    await db.tracks.delete_one({"id": track_id})
    
    return {"message": "Track and associated files deleted successfully"}

@api_router.get("/tracks/{track_id}", response_model=MusicTrack)
async def get_track_details(track_id: str, current_user: User = Depends(get_current_user)):
    """
    Fetch the details for a single music track.
    """
    track = await db.tracks.find_one({"id": track_id})
    if not track:
        raise HTTPException(status_code=404, detail="Track not found")
    
    # Authorization check: Managers can only view their own tracks
    if current_user.user_type == "manager":
        if track.get("created_by") != current_user.id and track.get("managed_by") != current_user.manager_id:
            raise HTTPException(status_code=403, detail="Not authorized to view this track")

    return MusicTrack(**parse_from_mongo(track))

@api_router.get("/tracks/{track_id}/download/{file_type}")
async def download_track_file(track_id: str, file_type: str, current_user: User = Depends(get_current_user)):
    """
    Generates a temporary signed URL for downloading a specific file and redirects to it.
    """
    track = await db.tracks.find_one({"id": track_id})
    if not track:
        raise HTTPException(status_code=404, detail="Track not found")

    # Authorization check
    if current_user.user_type == "manager":
        if track.get("created_by") != current_user.id and track.get("managed_by") != current_user.manager_id:
            raise HTTPException(status_code=403, detail="Not authorized to download this file")

    # Map the file type from the URL to the correct blob name field in the database
    blob_field_map = {
        "mp3": "mp3_blob_name",
        "lyrics": "lyrics_blob_name",
        "session": "session_blob_name",
        "singer_agreement": "singer_agreement_blob_name",
        "music_director_agreement": "music_director_agreement_blob_name"
    }
    
    blob_field = blob_field_map.get(file_type)
    if not blob_field or not track.get(blob_field):
        raise HTTPException(status_code=404, detail=f"File of type '{file_type}' not found for this track.")

    blob_name = track[blob_field]
    
    # Generate a signed URL for reading (GET) the file
    try:
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(blob_name)
        
        # Generate a URL that forces download by setting response-disposition
        original_filename = track.get(f"{file_type}_filename") or blob_name.split('/')[-1]
        
        url = blob.generate_signed_url(
            version="v4",
            expiration=timedelta(minutes=15),
            method="GET",
            response_disposition=f'attachment; filename="{original_filename}"'
        )
        # Redirect the user's browser directly to the GCS download link
        return RedirectResponse(url=url)
    except Exception as e:
        logger.exception(f"Failed to generate download URL for blob: {blob_name}")
        raise HTTPException(status_code=500, detail="Could not generate download link.")


@api_router.get("/tracks/{track_id}/stream")
async def stream_track_audio(track_id: str, current_user: User = Depends(get_current_user)):
    """
    Generates a temporary signed URL for streaming the audio file and redirects to it.
    """
    track = await db.tracks.find_one({"id": track_id})
    if not track:
        raise HTTPException(status_code=404, detail="Track not found")

    # Authorization check
    if current_user.user_type == "manager":
        if track.get("created_by") != current_user.id and track.get("managed_by") != current_user.manager_id:
            raise HTTPException(status_code=403, detail="Not authorized to stream this file")

    blob_name = track.get("mp3_blob_name")
    if not blob_name:
        raise HTTPException(status_code=404, detail="Audio file not found for this track.")
    
    # Generate a signed URL for reading (GET) the file
    try:
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(blob_name)
        url = blob.generate_signed_url(
            version="v4",
            expiration=timedelta(hours=2), # Longer expiration for audio streaming
            method="GET"
        )
        # Redirect the user's browser directly to the GCS stream link
        return RedirectResponse(url=url)
    except Exception as e:
        logger.exception(f"Failed to generate stream URL for blob: {blob_name}")
        raise HTTPException(status_code=500, detail="Could not generate audio stream link.")        

@api_router.get("/tracks/next-code/{full_prefix}")
async def get_next_unique_code(full_prefix: str, current_user: User = Depends(get_current_user)):
    """Generate the next available unique code for the given language-prefix combination"""
    
    # Parse the full prefix (e.g., "ENG-MR", "TEL-OCC", "HIN-OCW")
    if '-' not in full_prefix:
        raise HTTPException(status_code=400, detail="Invalid format. Expected format: LANG-PREFIX (e.g., ENG-MR)")
    
    parts = full_prefix.split('-', 1)
    if len(parts) != 2:
        raise HTTPException(status_code=400, detail="Invalid format. Expected format: LANG-PREFIX (e.g., ENG-MR)")
    
    language_code, prefix = parts
    
    # Validate language code (3 letters)
    if len(language_code) != 3 or not language_code.isalpha():
        raise HTTPException(status_code=400, detail="Language code must be 3 letters (e.g., ENG, TEL, HIN)")
    
    # Validate prefix
    valid_prefixes = ['OC', 'OCC', 'OCW', 'MR']
    if prefix not in valid_prefixes:
        raise HTTPException(status_code=400, detail=f"Invalid prefix. Must be one of: {valid_prefixes}")
    
    # Find the highest existing code with this full prefix (language + rights prefix)
    regex_pattern = f"^{language_code}-{prefix}\\d+$"
    last_track = await db.tracks.find_one(
        {"unique_code": {"$regex": regex_pattern}},
        sort=[("unique_code", -1)]
    )
    
    if last_track and last_track.get("unique_code"):
        # Extract number from last unique code
        try:
            last_code = last_track["unique_code"]
            # Remove language-prefix part and convert to int (e.g., "ENG-MR0003" -> "0003" -> 3)
            number_part = last_code[len(f"{language_code}-{prefix}"):]
            last_number = int(number_part)
            next_number = last_number + 1
        except (ValueError, IndexError):
            next_number = 1
    else:
        next_number = 1
    
    # Format the unique code with leading zeros (4 digits)
    unique_code = f"{language_code}-{prefix}{next_number:04d}"
    
    # Double check this code doesn't exist (in case of race conditions)
    existing_track = await db.tracks.find_one({"unique_code": unique_code})
    if existing_track:
        # If exists, try next number
        next_number += 1
        unique_code = f"{language_code}-{prefix}{next_number:04d}"
    logger.info(f"Generated unique code: {unique_code} for prefix: {full_prefix}")
    return {"unique_code": unique_code}

@api_router.put("/tracks/{track_id}", response_model=MusicTrack)
async def update_track(
    track_id: str,
    track_update: MusicTrackUpdate,
    current_user: User = Depends(get_current_user)
):
    track = await db.tracks.find_one({"id": track_id})
    if not track:
        raise HTTPException(status_code=404, detail="Track not found")
    
    # Authorization: Admins can edit any track, Managers can only edit their own tracks
    if current_user.user_type == "manager":
        if track["created_by"] != current_user.id and track.get("managed_by") != current_user.manager_id:
            raise HTTPException(status_code=403, detail="Not authorized to update this track")
    # Admins can edit any track (no additional check needed)
    
    update_data = {k: v for k, v in track_update.dict().items() if v is not None}
    
    if update_data:
        await db.tracks.update_one({"id": track_id}, {"$set": update_data})
    
    updated_track = await db.tracks.find_one({"id": track_id})
    return MusicTrack(**parse_from_mongo(updated_track))

@api_router.get("/verify-deployment")
async def verify_deployment():
    """A simple endpoint to confirm the latest code is deployed."""
    return {
        "message": "Deployment successful! The new code is live.",
        "version": "v3-library-update-check",
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

@api_router.get("/verify-signing")
async def verify_signing():
    """
    An advanced diagnostic endpoint to test the core IAM signing permission,
    bypassing the GCS library entirely.
    """
    try:
        logger.info("--- Starting core signing verification ---")
        
        # This is the service account your Cloud Run service uses
        target_service_account = os.environ.get('SIGNING_SERVICE_ACCOUNT_EMAIL')
        if not target_service_account:
            raise ValueError("SIGNING_SERVICE_ACCOUNT_EMAIL env var is not set.")

        logger.info(f"Attempting to sign using the identity of: {target_service_account}")
        
        # We create credentials that are "impersonating" our own service account.
        # This is the action that requires the 'Service Account Token Creator' role.
        creds = impersonated_credentials.Credentials(
            source_credentials=None,  # Use application default credentials
            target_principal=target_service_account,
            target_scopes=["https://www.googleapis.com/auth/devstorage.read_write"],
            # The signer is the IAM Credentials API, not a local private key
        )
        
        # The data we want to sign
        payload_to_sign = "test-payload-for-signing".encode('utf-8')

        # This is the direct call to the IAM Credentials API to sign a blob of bytes.
        # This is what the GCS library *should* be doing under the hood.
        signed_blob = await asyncio.to_thread(creds.sign_bytes, payload_to_sign)

        logger.info("--- Core signing verification SUCCEEDED ---")
        
        return {
            "status": "SUCCESS",
            "message": "The service account has the required permissions to sign data using the IAM Credentials API.",
            "signed_payload_base64": base64.b64encode(signed_blob).decode('utf-8')
        }

    except Exception as e:
        logger.exception("--- Core signing verification FAILED ---")
        return {
            "status": "FAILED",
            "error_type": type(e).__name__,
            "error_message": str(e),
            "instructions": "This failure indicates a fundamental permissions issue, likely from a Google Cloud Organization Policy. Please contact your GCP administrator with this error message."
        }

@api_router.get("/verify-libraries")
async def verify_libraries():
    """Endpoint to verify the installed versions of key libraries."""
    try:
        import google.cloud.storage
        import google.auth
        
        gcs_version = getattr(google.cloud.storage, "__version__", "Not Found")
        auth_version = getattr(google.auth, "__version__", "Not Found")
        
        return {
            "message": "Installed library versions",
            "google_cloud_storage_version": gcs_version,
            "google_auth_version": auth_version
        }
    except Exception as e:
        return {"error": str(e)}

@api_router.post("/tracks/generate-upload-url")
async def generate_upload_url(
    filename: str = Form(...),
    content_type: str = Form(...),
    folder: str = Form(...),
    current_user: User = Depends(get_current_user)
):
    """
    Generate a signed URL using explicit, manual impersonation to fix credential auto-discovery issues.
    This is the definitive and final version.
    """
    logger.info(f"Upload URL request from user {current_user.id}: folder={folder}, file={filename}")

    # --- Initial Configuration Checks ---
    if not GCS_BUCKET_NAME:
        logger.error("FATAL: GCS_BUCKET_NAME environment variable is not set.")
        raise HTTPException(status_code=500, detail="Server is misconfigured: Storage bucket name is missing.")
    
    if not SIGNING_SERVICE_ACCOUNT_EMAIL:
        logger.error("FATAL: SIGNING_SERVICE_ACCOUNT_EMAIL environment variable is not set.")
        raise HTTPException(status_code=500, detail="Server is misconfigured: Signing service account is missing.")

    # --- Input Validation ---
    valid_folders = ['audio', 'lyrics', 'sessions', 'agreements']
    if folder not in valid_folders:
        raise HTTPException(status_code=400, detail=f"Invalid folder specified.")
    
    # You can add your content-type validation logic here if needed
    await require_upload_permission(folder, current_user)

    try:
        # --- Prepare GCS Blob ---
        safe_filename = sanitize_filename(filename)
        blob_name = f"{folder}/{uuid.uuid4()}_{safe_filename}"
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(blob_name)

        logger.info("--- Starting Explicit Impersonation Process ---")

        # Step 1: Manually fetch the base credentials from the Cloud Run environment (ADC).
        # This is the key fix that prevents the 'NoneType' error by forcing the credential discovery.
        base_creds, project_id = google.auth.default(scopes=["https://www.googleapis.com/auth/cloud-platform"])
        if not project_id:
             logger.warning("Could not determine project ID from default credentials.")
        logger.info(f"Base credentials successfully obtained for project: {project_id}")

        # Step 2: Manually create the impersonated credentials object.
        # We explicitly provide the base credentials, bypassing the broken auto-discovery mechanism.
        impersonated_creds = impersonated_credentials.Credentials(
            source_credentials=base_creds,
            target_principal=SIGNING_SERVICE_ACCOUNT_EMAIL,
            target_scopes=["https://www.googleapis.com/auth/devstorage.read_write"],
            lifetime=900  # Lifetime in seconds (15 minutes), matching the URL expiration
        )
        logger.info(f"Impersonated credentials object successfully created for target: {SIGNING_SERVICE_ACCOUNT_EMAIL}")

        # Step 3: Generate the signed URL using the EXPLICIT credentials object.
        # We now pass the 'credentials=' parameter instead of 'service_account_email='.
        signed_url = await asyncio.to_thread(
            blob.generate_signed_url,
            expiration=timedelta(minutes=15),
            method="PUT",
            version="v4",
            content_type=content_type,
            credentials=impersonated_creds,  # <-- THE CRITICAL FIX IS HERE
        )
        
        logger.info("--- Explicit Impersonation and URL Signing SUCCEEDED ---")

    except Exception as e:
        logger.exception("CRITICAL ERROR during explicit signed URL generation")
        # Provide a detailed error message to the client for easier debugging.
        raise HTTPException(status_code=500, detail=f"Failed to generate upload URL: {type(e).__name__}: {str(e)}") from e
    
    return {
            "signed_url": signed_url,
            "blob_name": blob_name,
            "filename": filename
    }

@api_router.delete("/tracks/cleanup-upload/{blob_name:path}")
async def cleanup_upload(
    blob_name: str,
    current_user: User = Depends(get_current_user)
):
    """Delete an uploaded file from GCS storage"""
    if not blob_name:
        raise HTTPException(status_code=400, detail="Blob name is required")

    # Validate blob path structure
    valid_folders = ['audio', 'lyrics', 'sessions', 'agreements']
    folder = blob_name.split('/')[0] if '/' in blob_name else None
    
    if not folder or folder not in valid_folders:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid blob path. Must start with one of: {valid_folders}"
        )

    # Check user's permission for the folder
    await require_upload_permission(folder, current_user)

    try:
        # Attempt to delete the blob
        await delete_from_gcs(blob_name)
        logger.info(f"Successfully cleaned up blob {blob_name} for user {current_user.id}")
        return {"message": "File cleaned up successfully"}
        
    except NotFound:
        # File already deleted or doesn't exist
        raise HTTPException(
            status_code=404,
            detail="File not found"
        ) from None
        
    except Forbidden:
        # Permission issues with GCS
        raise HTTPException(
            status_code=403,
            detail="Permission denied to delete file"
        ) from None
        
    except Exception as e:
        # Log unexpected errors but don't expose details to client
        logger.exception("Failed to clean up blob", extra={"blob_name": blob_name})
        raise HTTPException(
            status_code=500,
            detail="Failed to clean up file"
        ) from None

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
